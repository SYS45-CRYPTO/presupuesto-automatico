import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from decimal import Decimal
from datetime import datetime
from typing import Dict, List, Any, Optional
import logging

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Exporta presupuestos y datos a formato Excel"""
    
    def __init__(self):
        self.workbook = None
        self.styles = {}
    
    def export_budget_to_excel(self, budget_data: Dict[str, Any], output_path: str, 
                             include_charts: bool = True) -> bool:
        """
        Exporta un presupuesto completo a Excel
        
        Args:
            budget_data: Datos completos del presupuesto
            output_path: Ruta de salida del archivo Excel
            include_charts: Si incluir gráficos
            
        Returns:
            True si se exportó exitosamente
        """
        try:
            logger.info(f"Exportando presupuesto a Excel: {output_path}")
            
            # Crear workbook
            self.workbook = Workbook()
            
            # Remover hoja por defecto
            self.workbook.remove(self.workbook.active)
            
            # Crear hojas
            self._create_summary_sheet(budget_data)
            self._create_detailed_items_sheet(budget_data)
            self._create_cost_analysis_sheet(budget_data)
            
            if include_charts:
                self._create_charts_sheet(budget_data)
            
            # Guardar archivo
            self.workbook.save(output_path)
            
            logger.info(f"Presupuesto exportado exitosamente: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exportando a Excel: {str(e)}")
            return False
    
    def export_multiple_budgets(self, budgets_data: List[Dict[str, Any]], 
                              output_path: str) -> bool:
        """
        Exporta múltiples presupuestos para comparación
        
        Args:
            budgets_data: Lista de presupuestos
            output_path: Ruta de salida
            
        Returns:
            True si se exportó exitosamente
        """
        try:
            logger.info(f"Exportando comparación de presupuestos: {output_path}")
            
            self.workbook = Workbook()
            self.workbook.remove(self.workbook.active)
            
            # Hoja de comparación
            self._create_comparison_sheet(budgets_data)
            
            # Hoja de análisis
            self._create_comparison_analysis_sheet(budgets_data)
            
            self.workbook.save(output_path)
            
            logger.info(f"Comparación exportada exitosamente: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exportando comparación: {str(e)}")
            return False
    
    def export_price_book(self, price_book_data: Dict[str, Any], output_path: str) -> bool:
        """
        Exporta un libro de precios a Excel
        
        Args:
            price_book_data: Datos del libro de precios
            output_path: Ruta de salida
            
        Returns:
            True si se exportó exitosamente
        """
        try:
            logger.info(f"Exportando libro de precios: {output_path}")
            
            self.workbook = Workbook()
            self.workbook.remove(self.workbook.active)
            
            # Hoja principal de precios
            self._create_price_book_sheet(price_book_data)
            
            # Hoja de estadísticas
            self._create_price_book_stats_sheet(price_book_data)
            
            # Hoja de histórico (si existe)
            if 'price_history' in price_book_data:
                self._create_price_history_sheet(price_book_data['price_history'])
            
            self.workbook.save(output_path)
            
            logger.info(f"Libro de precios exportado exitosamente: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exportando libro de precios: {str(e)}")
            return False
    
    def _create_summary_sheet(self, budget_data: Dict[str, Any]):
        """Crea hoja de resumen del presupuesto"""
        
        ws = self.workbook.create_sheet("Resumen")
        
        # Títulos principales
        ws['A1'] = "PRESUPUESTO DE OBRA"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Información del proyecto
        ws['A3'] = "Proyecto:"
        ws['B3'] = budget_data.get('project', {}).get('name', '')
        ws['A4'] = "Cliente:"
        ws['B4'] = budget_data.get('project', {}).get('client_name', '')
        ws['A5'] = "Ubicación:"
        ws['B5'] = budget_data.get('project', {}).get('location', '')
        ws['A6'] = "Fecha:"
        ws['B6'] = datetime.now().strftime('%d/%m/%Y')
        
        # Formato para etiquetas
        for row in range(3, 7):
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].alignment = Alignment(horizontal='right')
        
        # Resumen de costos
        ws['A8'] = "RESUMEN DE COSTOS"
        ws['A8'].font = Font(size=12, bold=True)
        ws['A8'].fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        ws.merge_cells('A8:C8')
        
        # Datos de costos
        cost_data = [
            ['Concepto', 'Valor', 'Porcentaje'],
            ['Costos Directos', budget_data.get('total_amount', 0) - budget_data.get('profit_amount', 0),
             ((budget_data.get('total_amount', 0) - budget_data.get('profit_amount', 0)) / budget_data.get('total_amount', 1) * 100)],
            ['Beneficio', budget_data.get('profit_amount', 0),
             (budget_data.get('profit_amount', 0) / budget_data.get('total_amount', 1) * 100)],
            ['TOTAL PRESUPUESTO', budget_data.get('final_amount', 0), 100.0]
        ]
        
        # Escribir datos de costos
        for i, (concept, value, percentage) in enumerate(cost_data):
            row = 9 + i
            ws[f'A{row}'] = concept
            ws[f'B{row}'] = value
            ws[f'C{row}'] = f"{percentage:.1f}%"
            
            # Formato de moneda para valores
            if i > 0:  # Excluir encabezados
                ws[f'B{row}'].number_format = '$#,##0.00'
            
            # Formato para totales
            if concept == 'TOTAL PRESUPUESTO':
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'C{row}'].font = Font(bold=True)
                ws[f'A{row}'].fill = PatternFill(start_color="D5E8F7", end_color="D5E8F7", fill_type="solid")
                ws[f'B{row}'].fill = PatternFill(start_color="D5E8F7", end_color="D5E8F7", fill_type="solid")
                ws[f'C{row}'].fill = PatternFill(start_color="D5E8F7", end_color="D5E8F7", fill_type="solid")
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        
        # Agregar bordes
        self._add_borders_to_range(ws, f'A9:C{9 + len(cost_data) - 1}')
    
    def _create_detailed_items_sheet(self, budget_data: Dict[str, Any]):
        """Crea hoja de detalle de partidas"""
        
        ws = self.workbook.create_sheet("Partidas Detalladas")
        
        # Encabezados
        headers = ['Capítulo', 'Código', 'Descripción', 'Unidad', 'Cantidad', 'P. Unitario', 'Total']
        for i, header in enumerate(headers):
            cell = ws.cell(row=1, column=i+1, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Agrupar items por capítulo
        chapters = self._group_items_by_chapter(budget_data.get('items', []))
        
        row = 2
        for chapter, data in chapters.items():
            # Escribir items del capítulo
            for item in data['items']:
                ws.cell(row=row, column=1, value=chapter)
                ws.cell(row=row, column=2, value=item.get('code', ''))
                ws.cell(row=row, column=3, value=item.get('description', ''))
                ws.cell(row=row, column=4, value=item.get('unit', ''))
                ws.cell(row=row, column=5, value=item.get('quantity', 0))
                ws.cell(row=row, column=6, value=item.get('unit_price', 0))
                ws.cell(row=row, column=7, value=item.get('total_price', 0))
                
                # Formato de moneda
                ws[f'F{row}'].number_format = '$#,##0.00'
                ws[f'G{row}'].number_format = '$#,##0.00'
                
                # Alinear números a la derecha
                ws[f'E{row}'].alignment = Alignment(horizontal='right')
                ws[f'F{row}'].alignment = Alignment(horizontal='right')
                ws[f'G{row}'].alignment = Alignment(horizontal='right')
                
                row += 1
            
            # Subtotal del capítulo
            ws.cell(row=row, column=6, value="SUBTOTAL")
            ws[f'F{row}'].font = Font(bold=True)
            ws.cell(row=row, column=7, value=data['total'])
            ws[f'G{row}'].font = Font(bold=True)
            ws[f'G{row}'].number_format = '$#,##0.00'
            ws[f'G{row}'].fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
            
            row += 2  # Espacio entre capítulos
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
    
    def _create_cost_analysis_sheet(self, budget_data: Dict[str, Any]):
        """Crea hoja de análisis de costos"""
        
        ws = self.workbook.create_sheet("Análisis de Costos")
        
        # Título
        ws['A1'] = "ANÁLISIS DE COSTOS"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # Desglose de costos
        cost_breakdown = budget_data.get('cost_breakdown', {})
        
        # Tabla de desglose
        breakdown_data = [
            ['COMPONENTE', 'VALOR', 'PORCENTAJE SOBRE TOTAL', ''],
            ['Mano de Obra', cost_breakdown.get('labor_cost', 0),
             (cost_breakdown.get('labor_cost', 0) / budget_data.get('total_amount', 1) * 100), ''],
            ['Materiales', cost_breakdown.get('material_cost', 0),
             (cost_breakdown.get('material_cost', 0) / budget_data.get('total_amount', 1) * 100), ''],
            ['Equipo y Maquinaria', cost_breakdown.get('equipment_cost', 0),
             (cost_breakdown.get('equipment_cost', 0) / budget_data.get('total_amount', 1) * 100), ''],
            ['Costos Indirectos', cost_breakdown.get('indirect_cost', 0),
             (cost_breakdown.get('indirect_cost', 0) / budget_data.get('total_amount', 1) * 100), ''],
            ['Beneficio', cost_breakdown.get('profit_amount', 0),
             (cost_breakdown.get('profit_amount', 0) / budget_data.get('total_amount', 1) * 100), ''],
            ['TOTAL', budget_data.get('total_amount', 0), 100.0, '']
        ]
        
        # Escribir datos
        for i, (component, value, percentage, _) in enumerate(breakdown_data):
            row = 3 + i
            ws[f'A{row}'] = component
            ws[f'B{row}'] = value
            ws[f'C{row}'] = f"{percentage:.1f}%"
            
            # Formato de moneda
            if i > 0 and component != 'TOTAL':
                ws[f'B{row}'].number_format = '$#,##0.00'
            
            # Formato para totales
            if component == 'TOTAL':
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'C{row}'].font = Font(bold=True)
                ws[f'A{row}'].fill = PatternFill(start_color="D5E8F7", end_color="D5E8F7", fill_type="solid")
                ws[f'B{row}'].fill = PatternFill(start_color="D5E8F7", end_color="D5E8F7", fill_type="solid")
                ws[f'C{row}'].fill = PatternFill(start_color="D5E8F7", end_color="D5E8F7", fill_type="solid")
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        # Agregar bordes
        self._add_borders_to_range(ws, f'A3:C{3 + len(breakdown_data) - 1}')
        
        # Notas adicionales
        notes_row = 3 + len(breakdown_data) + 2
        ws[f'A{notes_row}'] = "NOTAS:"
        ws[f'A{notes_row}'].font = Font(bold=True)
        
        notes = [
            "• Los precios incluyen todos los costos directos e indirectos necesarios para la ejecución de la obra.",
            "• Los precios no incluyen IVA.",
            "• La vigencia de este presupuesto es de 30 días calendario.",
            "• Cualquier modificación en el alcance de los trabajos deberá ser notificada por escrito."
        ]
        
        for i, note in enumerate(notes):
            ws[f'A{notes_row + 1 + i}'] = note
            ws[f'A{notes_row + 1 + i}'].alignment = Alignment(wrap_text=True)
    
    def _create_charts_sheet(self, budget_data: Dict[str, Any]):
        """Crea hoja con gráficos"""
        
        ws = self.workbook.create_sheet("Gráficos")
        
        # Gráfico de pastel para desglose de costos
        cost_breakdown = budget_data.get('cost_breakdown', {})
        
        # Datos para el gráfico
        chart_data = [
            ['Componente', 'Valor'],
            ['Mano de Obra', cost_breakdown.get('labor_cost', 0)],
            ['Materiales', cost_breakdown.get('material_cost', 0)],
            ['Equipo', cost_breakdown.get('equipment_cost', 0)],
            ['Costos Indirectos', cost_breakdown.get('indirect_cost', 0)],
            ['Beneficio', cost_breakdown.get('profit_amount', 0)]
        ]
        
        # Escribir datos del gráfico
        for i, (component, value) in enumerate(chart_data):
            ws[f'A{2 + i}'] = component
            ws[f'B{2 + i}'] = value
        
        # Crear gráfico de pastel
        pie = PieChart()
        pie.title = "Desglose de Costos"
        
        labels = Reference(ws, min_col=1, min_row=3, max_row=7)
        data = Reference(ws, min_col=2, min_row=2, max_row=7)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        
        # Colores personalizados
        colors = ["4472C4", "70AD47", "FFC000", "FF6D01", "C55A5A", "7030A0"]
        for i, color in enumerate(colors):
            if i < len(pie.series[0].points):
                pie.series[0].points[i].graphicalProperties.solidFill = color
        
        ws.add_chart(pie, "D2")
        
        # Gráfico de barras por capítulo
        chapters = self._group_items_by_chapter(budget_data.get('items', []))
        
        # Datos para gráfico de barras
        bar_data = [['Capítulo', 'Valor']]
        for chapter, data in chapters.items():
            bar_data.append([chapter, data['total']])
        
        # Escribir datos
        start_row = 10
        for i, (chapter, value) in enumerate(bar_data):
            ws[f'A{start_row + i}'] = chapter
            ws[f'B{start_row + i}'] = value
        
        # Crear gráfico de barras
        bar = BarChart()
        bar.title = "Costos por Capítulo"
        bar.y_axis.title = 'Valor ($)'
        
        labels = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(bar_data) - 1)
        data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(bar_data) - 1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(labels)
        
        ws.add_chart(bar, "D15")
    
    def _create_comparison_sheet(self, budgets_data: List[Dict[str, Any]]):
        """Crea hoja de comparación de presupuestos"""
        
        ws = self.workbook.create_sheet("Comparación")
        
        # Título
        ws['A1'] = "COMPARACIÓN DE PRESUPUESTOS"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Tabla comparativa
        headers = ['CONCEPTO'] + [budget.get('name', f"Presupuesto {i+1}") for i, budget in enumerate(budgets_data)]
        
        # Escribir encabezados
        for i, header in enumerate(headers):
            cell = ws.cell(row=3, column=i+1, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de comparación
        comparison_data = [
            ['Subtotal', 'total_amount - profit_amount'],
            ['Beneficio', 'profit_amount'],
            ['Total', 'final_amount'],
            ['Margen de Beneficio (%)', 'profit_margin'],
            ['Número de Partidas', 'item_count'],
            ['Fecha de Creación', 'created_at']
        ]
        
        for i, (concept, field) in enumerate(comparison_data):
            row = 4 + i
            ws[f'A{row}'] = concept
            ws[f'A{row}'].font = Font(bold=True)
            
            for j, budget in enumerate(budgets_data):
                col = j + 2
                
                if field == 'total_amount - profit_amount':
                    value = budget.get('total_amount', 0) - budget.get('profit_amount', 0)
                elif field == 'profit_amount':
                    value = budget.get('profit_amount', 0)
                elif field == 'final_amount':
                    value = budget.get('final_amount', 0)
                elif field == 'profit_margin':
                    total = budget.get('total_amount', 1)
                    profit = budget.get('profit_amount', 0)
                    value = (profit / total * 100) if total > 0 else 0
                    value = f"{value:.1f}%"
                elif field == 'item_count':
                    value = len(budget.get('items', []))
                elif field == 'created_at':
                    value = budget.get('created_at', 'N/A')
                else:
                    value = budget.get(field, '')
                
                ws.cell(row=row, column=col, value=value)
                
                # Formato de moneda para valores
                if field in ['total_amount - profit_amount', 'profit_amount', 'final_amount']:
                    ws.cell(row=row, column=col).number_format = '$#,##0.00'
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 25
        for i in range(len(budgets_data)):
            col_letter = chr(ord('B') + i)
            ws.column_dimensions[col_letter].width = 15
        
        # Agregar bordes
        self._add_borders_to_range(ws, f'A3:{chr(ord("A") + len(budgets_data))}{3 + len(comparison_data) - 1}')
    
    def _create_comparison_analysis_sheet(self, budgets_data: List[Dict[str, Any]]):
        """Crea hoja de análisis de comparación"""
        
        ws = self.workbook.create_sheet("Análisis Comparativo")
        
        if len(budgets_data) >= 2:
            budget1 = budgets_data[0]
            budget2 = budgets_data[1]
            
            ws['A1'] = "ANÁLISIS DE VARIACIONES"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:C1')
            
            # Calcular variaciones
            total_diff = budget2.get('final_amount', 0) - budget1.get('final_amount', 0)
            total_pct = (total_diff / budget1.get('final_amount', 1) * 100) if budget1.get('final_amount', 0) > 0 else 0
            
            items_diff = len(budget2.get('items', [])) - len(budget1.get('items', []))
            
            margin1 = (budget1.get('profit_amount', 0) / budget1.get('total_amount', 1) * 100)
            margin2 = (budget2.get('profit_amount', 0) / budget2.get('total_amount', 1) * 100)
            margin_diff = margin2 - margin1
            
            # Escribir análisis
            analysis_data = [
                ['Variación Total', f"${total_diff:,.2f}", f"{total_pct:.1f}%"],
                ['Diferencia en Partidas', items_diff, 'partidas'],
                ['Variación de Margen', f"{margin_diff:.1f}%", 'puntos'],
            ]
            
            for i, (concept, value, unit) in enumerate(analysis_data):
                row = 3 + i
                ws[f'A{row}'] = concept
                ws[f'B{row}'] = value
                ws[f'C{row}'] = unit
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            
            # Agregar bordes
            self._add_borders_to_range(ws, f'A3:C{3 + len(analysis_data) - 1}')
    
    def _create_price_book_sheet(self, price_book_data: Dict[str, Any]):
        """Crea hoja de libro de precios"""
        
        ws = self.workbook.create_sheet("Precios")
        
        # Información del libro de precios
        ws['A1'] = f"Libro de Precios: {price_book_data.get('name', '')}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Descripción: {price_book_data.get('description', '')}"
        ws['A2'].font = Font(size=10)
        ws.merge_cells('A2:F2')
        
        # Encabezados
        headers = ['Código', 'Descripción', 'Unidad', 'Precio Unitario', 'Categoría', 'Notas']
        for i, header in enumerate(headers):
            cell = ws.cell(row=4, column=i+1, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de precios
        entries = price_book_data.get('entries', [])
        for i, entry in enumerate(entries):
            row = 5 + i
            ws.cell(row=row, column=1, value=entry.get('code', ''))
            ws.cell(row=row, column=2, value=entry.get('description', ''))
            ws.cell(row=row, column=3, value=entry.get('unit', ''))
            ws.cell(row=row, column=4, value=entry.get('unit_price', 0))
            ws.cell(row=row, column=5, value=entry.get('category', ''))
            ws.cell(row=row, column=6, value=entry.get('notes', ''))
            
            # Formato de moneda
            ws[f'D{row}'].number_format = '$#,##0.00'
            ws[f'D{row}'].alignment = Alignment(horizontal='right')
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 25
        
        # Agregar bordes
        if entries:
            self._add_borders_to_range(ws, f'A4:F{4 + len(entries)}')
    
    def _create_price_book_stats_sheet(self, price_book_data: Dict[str, Any]):
        """Crea hoja de estadísticas del libro de precios"""
        
        ws = self.workbook.create_sheet("Estadísticas")
        
        # Título
        ws['A1'] = "ESTADÍSTICAS DEL LIBRO DE PRECIOS"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        
        # Calcular estadísticas
        entries = price_book_data.get('entries', [])
        
        if entries:
            prices = [entry.get('unit_price', 0) for entry in entries]
            categories = list(set(entry.get('category', 'General') for entry in entries))
            units = list(set(entry.get('unit', 'un') for entry in entries))
            
            stats_data = [
                ['Total de Partidas', len(entries)],
                ['Precio Promedio', f"${sum(prices) / len(prices):,.2f}"],
                ['Precio Mínimo', f"${min(prices):,.2f}"],
                ['Precio Máximo', f"${max(prices):,.2f}"],
                ['Número de Categorías', len(categories)],
                ['Número de Unidades', len(units)]
            ]
            
            # Escribir estadísticas
            for i, (stat, value) in enumerate(stats_data):
                row = 3 + i
                ws[f'A{row}'] = stat
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_price_history_sheet(self, price_history: List[Dict[str, Any]]):
        """Crea hoja de histórico de precios"""
        
        ws = self.workbook.create_sheet("Histórico")
        
        # Encabezados
        headers = ['Código', 'Descripción', 'Precio Anterior', 'Precio Nuevo', 'Cambio', 'Fecha', 'Motivo']
        for i, header in enumerate(headers):
            cell = ws.cell(row=1, column=i+1, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        
        # Datos de histórico
        for i, history_entry in enumerate(price_history):
            row = 2 + i
            ws.cell(row=row, column=1, value=history_entry.get('code', ''))
            ws.cell(row=row, column=2, value=history_entry.get('description', ''))
            ws.cell(row=row, column=3, value=history_entry.get('previous_price', 0))
            ws.cell(row=row, column=4, value=history_entry.get('new_price', 0))
            
            # Calcular cambio
            prev_price = history_entry.get('previous_price', 0)
            new_price = history_entry.get('new_price', 0)
            change = new_price - prev_price
            change_pct = (change / prev_price * 100) if prev_price > 0 else 0
            ws.cell(row=row, column=5, value=f"${change:,.2f} ({change_pct:+.1f}%)")
            
            ws.cell(row=row, column=6, value=history_entry.get('change_date', ''))
            ws.cell(row=row, column=7, value=history_entry.get('change_reason', ''))
            
            # Formato de moneda
            ws[f'C{row}'].number_format = '$#,##0.00'
            ws[f'D{row}'].number_format = '$#,##0.00'
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30
        
        # Agregar bordes
        if price_history:
            self._add_borders_to_range(ws, f'A1:G{1 + len(price_history)}')
    
    def _group_items_by_chapter(self, items: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Agrupa items por capítulo"""
        
        chapters = {}
        
        for item in items:
            chapter = item.get('chapter', 'Sin Capítulo')
            
            if chapter not in chapters:
                chapters[chapter] = {
                    'description': f"Capítulo {chapter}",
                    'items': [],
                    'total': 0.0
                }
            
            chapters[chapter]['items'].append(item)
            chapters[chapter]['total'] += float(item.get('total_price', 0))
        
        return chapters
    
    def _add_borders_to_range(self, worksheet, range_string: str):
        """Agrega bordes a un rango de celdas"""
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet[range_string]:
            for cell in row:
                cell.border = thin_border